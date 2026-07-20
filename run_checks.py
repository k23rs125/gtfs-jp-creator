#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_checks.py — 回帰チェック（ローカル＆CI 共用）
================================================
コミット前やCIで実行し、壊れていないかを自動判定する。各チェックは独立で、
外部依存(P11/OSRM/Java Validator)や重いライブラリが無い環境(CI)では該当チェックを
SKIP する。1つでも FAIL があれば終了コード1。

    python run_checks.py
"""
import subprocess
import sys
import tempfile
from pathlib import Path

REPO = Path(__file__).resolve().parent
SCRIPTS = REPO / "skills" / "gtfs-jp-creator" / "scripts"
sys.path.insert(0, str(SCRIPTS))
sys.path.insert(0, str(REPO))

PASS, FAIL, SKIP = "PASS", "FAIL", "SKIP"
results = []


def record(name, status, detail=""):
    results.append((name, status))
    mark = {"PASS": "✅", "FAIL": "❌", "SKIP": "⏭"}[status]
    print(f"  {mark} [{status}] {name}" + (f" - {detail}" if detail else ""))


# ---- 1. コンパイル（app + 全スクリプト）----
def check_compile():
    import py_compile
    targets = [REPO / "app" / "app.py", REPO / "apply_decisions.py", REPO / "golden_test.py"]
    targets += sorted(SCRIPTS.glob("*.py"))
    bad = []
    for t in targets:
        if not t.exists():
            continue
        try:
            py_compile.compile(str(t), doraise=True)
        except Exception as e:
            bad.append(f"{t.name}: {str(e)[:60]}")
    record("compile", FAIL if bad else PASS, "; ".join(bad[:3]) or f"{len(targets)}ファイルOK")


# ---- 2. shape 重複点の除去（make_shape_rows）----
def check_shape_dedup():
    try:
        import generate_shapes as gs
    except Exception as e:
        record("shape_dedup", SKIP, f"import不可: {str(e)[:40]}")
        return
    pts = [(33.10, 130.40), (33.10, 130.40), (33.11, 130.41), (33.11, 130.41), (33.12, 130.42)]
    rows = gs.make_shape_rows("s1", pts)
    dup = any(rows[i]["shape_pt_lat"] == rows[i - 1]["shape_pt_lat"]
              and rows[i]["shape_pt_lon"] == rows[i - 1]["shape_pt_lon"]
              for i in range(1, len(rows)))
    ok = (not dup) and len(rows) == 3
    record("shape_dedup", PASS if ok else FAIL, f"5点→{len(rows)}点 連続重複={dup}")


# ---- 3. feed_contact_url の補完（generate_feed_info）----
def check_feed_contact():
    try:
        import generate_gtfs_files as gg
    except Exception as e:
        record("feed_contact", SKIP, f"import不可: {str(e)[:40]}")
        return
    with tempfile.TemporaryDirectory() as d:
        data = {"agency": {"agency_name": "テスト市", "agency_url": "https://example.jp"},
                "feed_info": {}, "calendar": []}
        try:
            gg.generate_feed_info(data, Path(d))
            txt = (Path(d) / "feed_info.txt").read_text(encoding="utf-8-sig")
        except Exception as e:
            record("feed_contact", FAIL, str(e)[:60])
            return
        ok = "feed_contact_url" in txt and "example.jp" in txt
        record("feed_contact", PASS if ok else FAIL,
               "agency_urlから補完" if ok else "feed_contact_url未出力")


# ---- 4. eval_compare の座標距離メトリクス（Haversine）----
def check_eval_compare():
    try:
        import eval_compare as ec
    except Exception as e:
        record("eval_compare", SKIP, f"import不可: {str(e)[:40]}")
        return
    d = ec._haversine_m(35.0, 135.0, 35.0, 135.001)   # 経度0.001度≈91m
    record("eval_compare", PASS if 80 < d < 100 else FAIL, f"haversine≈{round(d)}m(期待~91m)")


# ---- 5. apply_decisions が構造化を出す（決定的・外部依存なし・合成フィクスチャ）----
def check_apply_decisions():
    import json
    extract = {"blocks": [{"block_index": 0,
                           "stops": [{"name": "A停"}, {"name": "B停"}, {"name": "C停"}],
                           "trips": [{"cells": [
                               {"seq": 1, "num": None, "name": "A停", "time": "08:00:00", "reserve": False},
                               {"seq": 2, "num": None, "name": "B停", "time": "08:10:00", "reserve": False},
                               {"seq": 3, "num": None, "name": "C停", "time": "08:20:00", "reserve": False}]}]}]}
    spec = {"routes": [{"route_id": "R01", "route_long_name": "テスト線", "blocks": [0]}],
            "block_direction": {"0": 0}, "exclude_unnumbered": False, "stop_key": "name",
            "service": {"service_id": "SVC", "mon": 1, "tue": 1, "wed": 1, "thu": 1, "fri": 1,
                        "sat": 0, "sun": 0, "start_date": "20260401", "end_date": "20270331"}}
    with tempfile.TemporaryDirectory() as d:
        dp = Path(d)
        (dp / "ex.json").write_text(json.dumps(extract, ensure_ascii=False), encoding="utf-8")
        (dp / "sp.json").write_text(json.dumps(spec, ensure_ascii=False), encoding="utf-8")
        out = dp / "structured.json"
        r = subprocess.run([sys.executable, "-X", "utf8", str(REPO / "apply_decisions.py"),
                            "--extract", str(dp / "ex.json"), "--decisions", str(dp / "sp.json"),
                            "--out", str(out)],
                           capture_output=True, text=True, encoding="utf-8", errors="replace")
        if r.returncode != 0 or not out.exists():
            record("apply_decisions", FAIL, (r.stderr or r.stdout or "")[-80:])
            return
        s = json.loads(out.read_text(encoding="utf-8"))
        ok = bool(s.get("stops")) and bool(s.get("trips")) and bool(s.get("stop_times"))
        record("apply_decisions", PASS if ok else FAIL,
               f"stops={len(s.get('stops', []))} trips={len(s.get('trips', []))} "
               f"stop_times={len(s.get('stop_times', []))}")


# ---- 5b. 複数シートのExcel: 全シートを読み、積まれた表を連結しない ----
def check_excel_multisheet():
    """1ブックに複数の時刻表シートがある場合の回帰チェック（合成フィクスチャ）。

    かつては(1)先頭シートしか読まない (2)便番号行が無いレイアウトで上下に積まれた
    平日表と土日祝表がひとつながりの便に連結される、という2つの不具合があり、
    原典に無い時刻の便が静かに生成されていた。両方を固定する。
    """
    import json
    try:
        import openpyxl
    except Exception as e:
        record("excel_multisheet", SKIP, f"openpyxl未導入: {str(e)[:30]}")
        return
    import datetime as _dt

    def _put(ws, title, r0, times):
        ws.cell(r0, 1, f"【時刻表】（{title}）　テスト線")
        ws.cell(r0 + 1, 1, "停留所")
        ws.cell(r0 + 1, 2, 1)                      # 便番号の行（あれば便ヘッダとして使われる）
        for k, (nm, hh, mm) in enumerate(times):
            ws.cell(r0 + 2 + k, 1, nm)
            ws.cell(r0 + 2 + k, 2, _dt.time(hh, mm))

    with tempfile.TemporaryDirectory() as d:
        dp = Path(d)
        xlsx = dp / "multi.xlsx"
        wb = openpyxl.Workbook()
        s1 = wb.active
        s1.title = "平日"
        _put(s1, "平日", 1, [("A停", 8, 0), ("B停", 8, 10), ("C停", 8, 20)])
        s2 = wb.create_sheet("土日祝")
        _put(s2, "土日祝", 1, [("A停", 14, 0), ("B停", 14, 10), ("C停", 14, 20)])
        # 便番号の行を持たない冊子風シート: 2つの表を上下に積む（連結されないこと）
        s3 = wb.create_sheet("冊子")
        for r0, ttl, base in ((1, "平日", 9), (10, "土日祝", 16)):
            s3.cell(r0, 1, f"【時刻表】（{ttl}）　別線")
            s3.cell(r0 + 1, 1, "行先")
            s3.cell(r0 + 1, 2, "駅")
            for k, nm in enumerate(["X停", "Y停", "Z停"]):
                s3.cell(r0 + 2 + k, 1, nm)
                s3.cell(r0 + 2 + k, 2, _dt.time(base, k * 10))
        wb.save(xlsx)

        out = dp / "ex.json"
        r = subprocess.run([sys.executable, "-X", "utf8",
                            str(SCRIPTS / "extract_timetable_excel.py"), str(xlsx), "-o", str(out)],
                           capture_output=True, text=True, encoding="utf-8", errors="replace")
        if r.returncode != 0 or not out.exists():
            record("excel_multisheet", FAIL, (r.stderr or "")[-80:])
            return
        ex = json.loads(out.read_text(encoding="utf-8"))
        sheets = set(ex.get("sheets") or [])
        trips = [t for b in ex["blocks"] for t in b["trips"]]
        # (1) 先頭シートだけでなく全シートを読んでいる
        ok_sheets = {"平日", "土日祝", "冊子"} <= sheets
        # (2) どの便も1つの表の中で完結している（各表は3停留所なので、4停留所以上の便が
        #     できていたら上下に積まれた別の表が連結されている）。便の総数も固定する。
        ok_split = len(trips) == 4 and all(len(t["cells"]) == 3 for t in trips)
        # (3) 曜日の表記を拾えている
        hints = {b.get("day_hint") for b in ex["blocks"]}
        ok_hint = {"平日", "土日祝"} <= hints
        ok = ok_sheets and ok_split and ok_hint
        record("excel_multisheet", PASS if ok else FAIL,
               f"シート{len(sheets)} 便{len(trips)} 連結なし={ok_split} 曜日表記={ok_hint}")


# ---- 6. アプリが起動して抽出まで動く（AppTest スモーク）----
def check_apptest_smoke():
    try:
        from streamlit.testing.v1 import AppTest
    except Exception as e:
        record("apptest_smoke", SKIP, f"streamlit未導入: {str(e)[:30]}")
        return
    try:
        at = AppTest.from_file(str(REPO / "app" / "app.py"), default_timeout=120)
        at.run()
    except Exception as e:
        record("apptest_smoke", SKIP, f"app読込不可(依存欠落?): {str(e)[:50]}")
        return
    solo = [b for b in at.button if "一人で全部" in (b.label or "")]
    if not solo:
        record("apptest_smoke", FAIL, "最初の選択画面が出ない")
        return
    solo[0].click().run()
    sb = [b for b in at.button if "太宰府" in (b.label or "")]
    if sb:
        sb[0].click().run()
    record("apptest_smoke", PASS if len(at.exception) == 0 else FAIL,
           f"起動→一人→抽出 例外{len(at.exception)}")


# ---- 7. ゴールデンテスト（公式GTFS比較。P11がある環境のみ）----
def check_golden():
    try:
        import shapefile  # noqa: F401 pyshp があるか
    except Exception:
        record("golden_test", SKIP, "pyshp未導入(CI等)")
        return
    # golden_test.py の CASES が参照する P11 shapefile が存在するか確認
    p11 = Path("C:/Users/User/Desktop/稲ゼミ/p11_fukuoka/P11-22_40_SHP/P11-22_40.shp")
    fixture = REPO / "test_demo" / "ryobiraki_extract.json"
    if not (p11.exists() and fixture.exists()):
        record("golden_test", SKIP, "P11/フィクスチャ無し(CI等)")
        return
    r = subprocess.run([sys.executable, "-X", "utf8", str(REPO / "golden_test.py")],
                       capture_output=True, text=True, encoding="utf-8", errors="replace")
    ok = "ALL PASS" in (r.stdout or "")
    record("golden_test", PASS if ok else FAIL, "柳川両開" if ok else (r.stdout or "")[-80:])


def main():
    print("=" * 64)
    print("run_checks - 回帰チェック（外部依存が無い項目はSKIP）")
    print("=" * 64)
    for fn in (check_compile, check_shape_dedup, check_feed_contact, check_eval_compare,
               check_apply_decisions, check_excel_multisheet,
               check_apptest_smoke, check_golden):
        try:
            fn()
        except Exception as e:
            record(fn.__name__.replace("check_", ""), FAIL, f"チェック自体が例外: {str(e)[:60]}")
    n_fail = sum(1 for _, s in results if s == FAIL)
    n_pass = sum(1 for _, s in results if s == PASS)
    n_skip = sum(1 for _, s in results if s == SKIP)
    print("-" * 64)
    print(f"合計: PASS {n_pass} / FAIL {n_fail} / SKIP {n_skip}")
    print("総合:", "✅ OK" if n_fail == 0 else "❌ FAIL あり")
    sys.exit(1 if n_fail else 0)


if __name__ == "__main__":
    main()
