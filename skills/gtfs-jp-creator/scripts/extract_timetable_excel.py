#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract_timetable_excel.py
==========================
Step1 (Excel経路): Excel(.xlsx)のバス時刻表から停留所名・時刻を抽出し、
座標方式(extract_timetable_coords.py)と同じ「抽出JSON」形式で出力する。
以降の Step2(構造化)・Step3〜7(run_pipeline) は PDF と完全に共通化できる。

なぜ Excel を直接読むか:
  Excel は機械可読(セルが既にグリッド構造)なので、PDFの座標クラスタリングやOCRが
  不要で、最も確実・高精度に時刻表を読める。元データが Excel/Office にあるなら、
  PDF化せずそのまま読むのが「利用者の負担を減らし正確なデータを作る」設計に合う。

想定する標準レイアウト(よくあるバス時刻表):
  - 1列が「停留所名」（縦に停留所が並ぶ）
  - その右の複数列が「便」（各列=1便、セルが時刻 HH:MM）
  - 便名/行先などの見出しは時刻行の上にあってよい(任意)
  自動検出するが、外れる場合は --name-col / --header-rows / --sheet で上書き可。

設計方針(正しく失敗する):
  - 機械的に決められること(停留所名・時刻・便の並び)だけを行う。
  - 便名・方向・循環の解釈は行わない(Step2のLLM判断に委ねる)。
  - 構造が読めない/曖昧なときは推測せず warnings/needs_confirmation に明記。

Usage:
  python extract_timetable_excel.py <input.xlsx> -o <out.json>
        [--sheet <name>] [--name-col <Aや3>] [--header-rows <N>]

License: Apache 2.0
"""
import argparse
import datetime
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl が必要です (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)

TIME_RE = re.compile(r'^\s*(\d{1,2}):(\d{2})(?::\d{2})?\s*$')
# 先頭の空白(全角・半角)だけを落とす。かつて 店/駅 も剥がしていたが、
# 「駅前」「〇〇店前」等の実在停留所名を壊すため除外(正しいデータを最優先)。
ICON_PREFIX = re.compile(r'^[　\s]+')
DIRECTION_SUFFIX = re.compile(r'[（(](?:[東西南北]行き?|上り|下り|のりば\d*|\d+番のりば)[）)]\s*$')


def normalize_name(s: str) -> str:
    s = ICON_PREFIX.sub('', str(s)).strip()
    s = DIRECTION_SUFFIX.sub('', s).strip()
    return s


def route_title_from_texts(texts):
    """文字列群（タイトル行など）から路線名「○○線/系統/ルート」を1つ返す。
    曜日付記・付随語を落とし、全文字二重描画を畳む。JR等の鉄道路線・「系統番号」は除外。"""
    for text in texts:
        if not text:
            continue
        for tok in re.split(r"[\s　（）()【】\[\]／/｜|,、。「」]+", str(text)):
            tok = tok.strip()
            if len(tok) >= 4 and len(tok) % 2 == 0 and tok[0::2] == tok[1::2]:
                tok = tok[0::2]
            tok = re.sub(r"(時刻表|時刻|ダイヤ|運行表|一覧表|表)$", "", tok)
            if not (2 <= len(tok) <= 20):
                continue
            if any(x in tok for x in ("新幹線", "ゆたか線", "福北", "鉄道", "番号", "種類")):
                continue   # 鉄道路線・「系統番号」等は除外（JR古賀線等のバス路線名は許可）
            if tok.endswith(("線", "系統", "ルート")):
                return tok
    return None


# 時刻表のタイトルに書かれた運行日の表記。書かれている語だけを拾い、無ければ None
# （＝推測しない）。アプリ側で②の曜日チェックの初期値に使い、必ず要確認として出す。
DAY_HINT_PATTERNS = [
    ("平日", re.compile(r"平\s*日")),
    ("土日祝", re.compile(r"土\s*日\s*祝|土日・祝")),
    ("土曜", re.compile(r"土\s*曜")),
    ("日祝", re.compile(r"日\s*曜\s*[・､、]?\s*祝|日\s*祝")),
    ("毎日", re.compile(r"毎\s*日")),
]


def day_hint_from_texts(texts):
    """タイトル行の文字列から「平日/土日祝/土曜/日祝/毎日」の表記を1つ返す。無ければ None。"""
    for text in texts:
        if not text:
            continue
        s = str(text)
        for label, pat in DAY_HINT_PATTERNS:
            if pat.search(s):
                return label
    return None


def cell_time(v):
    """セル値を 'H:MM:00' に正規化。時刻でなければ None。"""
    if isinstance(v, datetime.datetime):
        return f"{v.hour}:{v.minute:02d}:00"
    if isinstance(v, datetime.time):
        return f"{v.hour}:{v.minute:02d}:00"
    if isinstance(v, str):
        m = TIME_RE.match(v)
        if m:
            return f"{int(m.group(1))}:{m.group(2)}:00"
    return None


def is_name_cell(v) -> bool:
    """停留所名らしいセルか（文字列・時刻でない・純数値でない）。"""
    if not isinstance(v, str):
        return False
    s = v.strip()
    if not s or cell_time(s) or s.isdigit():
        return False
    return True


def col_letter_to_idx(s: str) -> int:
    """'A'->1, 'B'->2 ... 数字ならそのまま int。"""
    s = s.strip()
    if s.isdigit():
        return int(s)
    n = 0
    for ch in s.upper():
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


def extract_sheet(ws, name_col_arg=None):
    """ワークシート1枚を抽出して {blocks, warnings, needs} を返す。

    複数シートのブックでも各シートを同じ手順で処理できるよう、main() から切り出したもの。
    blocks の block_index はこの関数内では 0 始まり。main() 側で通し番号に振り直す。
    """
    # --- 全セルを読む ---
    cells = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and str(c.value).strip() != "":
                cells[(c.row, c.column)] = c.value
    if not cells:
        return {"blocks": [], "warnings": [f"シート「{ws.title}」は空です。"], "needs": []}

    # --- スケジュール時刻セルと便(列) ---
    # 時刻型セルのうち「実時刻(hour>=4)」だけを便の時刻とみなす。所要時間列
    # (例: I/J/L/M列の 0:01, 0:05 など hour=0)を便の時刻と誤検出しないため。
    sched_cells = {}
    for (r, c), v in cells.items():
        t = cell_time(v)
        if t and int(t.split(":")[0]) >= 4:
            sched_cells[(r, c)] = t
    if not sched_cells:
        # 表紙・案内・運賃表など時刻表でないシートはここに来る。ブックに時刻表シートが
        # 他にあれば正常なので、エラーではなく「時刻表なし」として返す。
        return {"blocks": [], "warnings": [], "needs": [], "no_time_cells": True}

    col_sched_count = Counter(c for (r, c) in sched_cells)
    trip_cols = sorted([c for c, n in col_sched_count.items() if n >= 2])
    if not trip_cols:
        trip_cols = sorted(col_sched_count)  # 各列1便しか無い場合も拾う

    # --- 停留所名の列（横並びの複数表に対応） ---
    # 1シートに複数の独立した時刻表が横並びの場合（例: 往路/復路/行き/帰りがそれぞれ別の
    # 停留所名列を持つ）に対応。停留所名列を複数検出し、各名前列に「その右隣〜次の名前列
    # より左」の便列だけを対応付ける。単一表なら従来どおり1列になる。
    name_count = Counter(c for (r, c), v in cells.items() if is_name_cell(v))
    if name_col_arg:
        name_cols = [col_letter_to_idx(name_col_arg)]
    elif name_count:
        top = max(name_count.values())
        thr = max(4, int(top * 0.5))   # 表らしい名前列（散発的な少数セルは除外）
        name_cols = sorted(c for c, n in name_count.items() if n >= thr)
        if not name_cols:
            name_cols = [name_count.most_common(1)[0][0]]
    else:
        name_cols = []

    # 各名前列に便列を割り当て（その列より右、次の名前列より左）。便列が無い名前列
    # （例: 休憩時間表は実時刻が無い）は領域にしない＝自然に除外される。
    # band_lo は方向見出しの探索左端＝直前に採用した領域の右端（最初は1列目）。
    # これで方向ラベルが名前列の左にある場合（太宰府: ラベル列+停留所列）も拾え、
    # かつ横並び表で他領域の見出しに侵食しない。
    regions = []   # [(name_col, [trip_cols], band_lo, col_hi), ...]
    prev_hi = 1
    for i, nc in enumerate(name_cols):
        col_hi = name_cols[i + 1] if i + 1 < len(name_cols) else (ws.max_column + 1)
        rcols = [tc for tc in trip_cols if nc < tc < col_hi]
        if rcols:
            regions.append((nc, rcols, prev_hi, col_hi))
            prev_hi = col_hi
    if not regions and name_cols:   # フォールバック: 単一表として全便列を当てる
        regions = [(name_cols[0], trip_cols, 1, ws.max_column + 1)]

    warnings_list = []

    # --- セクション(縦表)検出: 便ヘッダ行で分割 ---
    # 便列の位置に便ラベル(便番号 or 「N便」文字列 or 上り/下り)が並び、実時刻が無い行を
    # 「便ヘッダ行」とみなす。その行を境に、同じ列を共有して縦に積まれた別方向の表
    # (例: 上り/下り、「市役所行き」「○○行き」)を分割する。
    # 便番号が「1便」のような文字列でも検出する（整数のみだと太宰府まほろば号で2方向を
    # 1つに連結してしまう取りこぼしがあった）。
    def _is_trip_label(v):
        if isinstance(v, bool):
            return False
        if isinstance(v, int):
            return 0 < v < 1000
        if isinstance(v, str):
            s = v.strip()
            if s.isdigit():
                return 0 < int(s) < 1000
            if re.match(r'^第?\s*\d+\s*便$', s):        # 「1便」「第1便」「10 便」
                return True
            if re.match(r'^[A-Za-zＡ-Ｚａ-ｚ]+\s*\d+\s*便$', s):   # 「B1便」「A3便」
                return True
            if s in ("上り", "下り", "往", "復", "往路", "復路", "行き", "帰り"):
                return True
        return False

    # 方向見出し（「市役所行き」「○○方面」「左回り/右回り」等）。便ヘッダの1つ上の行に
    # あることが多い。Step2 の方向/行先の手がかりとしてブロックに保持する。
    DIR_LABEL_RE = re.compile(r'(?:行き|帰り|方面|循環|回り|往路|復路)\s*$')

    # JR連絡（鉄道の発着）行はバス停ではないので停留所に含めない。
    def _is_jr(nm):
        return "JR" in nm.upper() or "ＪＲ" in nm

    def _direction_hint(hdr_row, col_lo, col_hi):
        # 見出しは領域内(col_lo〜col_hi)に限定して探す（横並び表で他領域の見出しを
        # 拾わないため）。便ヘッダの1〜2行上に「往路/復路/○○行き」等があることが多い。
        if hdr_row is None:
            return None
        # 見出しはヘッダの上(往路/復路等)にあることが多いが、ヘッダ行内(行き/帰り等)の
        # こともあるので上→ヘッダ行の順に探す。
        for rr in (hdr_row - 1, hdr_row - 2, hdr_row):
            for cc in range(col_lo, col_hi):
                v = cells.get((rr, cc))
                if isinstance(v, str) and DIR_LABEL_RE.search(v.strip()):
                    return v.strip()
        return None

    header_rows = []
    for r in range(1, ws.max_row + 1):
        labels = sum(1 for tc in trip_cols
                     if _is_trip_label(cells.get((r, tc))) and (r, tc) not in sched_cells)
        has_sched = any((r, tc) in sched_cells for tc in trip_cols)
        if labels >= 2 and not has_sched:
            header_rows.append(r)
    header_rows.sort()

    # タイトル行（「【時刻表】（平日）○○線」等）も表の切れ目とみなす。
    # 便番号の行が無い冊子用レイアウトでは便ヘッダ行を検出できず、上下に積まれた
    # 平日表と土日祝表がひとつながりの便として連結されてしまう（7:50発の便が
    # そのまま翌の9:45発へ続く、といった実在しない便になる）。タイトル行で切って防ぐ。
    title_rows = [r for r in range(1, ws.max_row + 1)
                  if any(isinstance(cells.get((r, c)), str) and "時刻表" in cells[(r, c)]
                         for c in range(1, ws.max_column + 1))]

    # 各セクションの範囲: (便ヘッダ行, 開始行, 終了行)。切れ目が無ければ単一表。
    _bnds = sorted(set(header_rows) | set(title_rows))
    if _bnds:
        _edges = _bnds + [ws.max_row + 1]
        sections = []
        for _i, _st in enumerate(_bnds):
            _end = _edges[_i + 1]
            # 便番号を読む行は、その区間の中にある便ヘッダ行（無ければ None）
            _hdr = next((h for h in header_rows if _st <= h < _end), None)
            sections.append((_hdr, _st, _end))
    else:
        sections = [(None, 0, ws.max_row + 1)]

    def _m(t):
        h, m, *_ = t.split(":")
        return int(h) * 60 + int(m)

    blocks = []
    needs = []
    all_served = set()
    for nc, rcols, band_lo, col_hi in regions:
        # この名前列の停留所行（JR連絡行は除外）
        row_name = {r: normalize_name(v) for (r, c), v in cells.items()
                    if c == nc and is_name_cell(v) and not _is_jr(normalize_name(v))}
        # 乗降制約マーカー（領域内のセルに「降車専用/乗車専用」）が付く行。範囲は推測せず、
        # マーカー行の停留所にヒントを付けて提示する（確定は人＝正しく失敗）。
        drop_rows = {r for (r, c), v in cells.items()
                     if nc <= c < col_hi and isinstance(v, str) and "降車専用" in v}
        pick_rows = {r for (r, c), v in cells.items()
                     if nc <= c < col_hi and isinstance(v, str) and "乗車専用" in v}
        for hdr, r_lo, r_hi in sections:
            sec_rows = sorted(r for r in row_name
                              if r_lo <= r < r_hi and (hdr is None or r > hdr))
            trips = []
            for tc in rcols:
                cell_list = []
                seq = 0
                for r in sec_rows:
                    if (r, tc) in sched_cells:
                        seq += 1
                        nm = row_name[r]
                        cell_list.append({"seq": seq, "num": None, "name": nm,
                                          "time": sched_cells[(r, tc)],
                                          "reserve": "要予約" in nm})
                if cell_list:
                    mins = [_m(c["time"]) for c in cell_list]
                    mono = all(mins[i] <= mins[i + 1] for i in range(len(mins) - 1))
                    tnum = cells.get((hdr, tc)) if hdr is not None else None
                    trips.append({"col": tc,
                                  "trip_number": str(tnum) if tnum is not None else None,
                                  "n_stops": len(cell_list), "monotonic": mono,
                                  "cells": cell_list})
            served = sorted({r for r in sec_rows
                             if any((r, tc) in sched_cells for tc in rcols)})
            all_served |= set(served)
            if not trips:
                continue
            stops = []
            for r in served:
                s = {"num": None, "name": row_name[r], "row": r,
                     "reserve": "要予約" in row_name[r]}
                if r in drop_rows:
                    s["boarding_hint"] = "drop_off_only"   # 降車専用（乗車不可）の疑い
                elif r in pick_rows:
                    s["boarding_hint"] = "pickup_only"      # 乗車専用（降車不可）の疑い
                stops.append(s)
            bi = len(blocks)
            blocks.append({"block_index": bi, "name_col": nc, "section_row": hdr,
                           "direction_hint": _direction_hint(hdr, band_lo, col_hi),
                           "stops": stops, "trips": trips, "warnings": []})
            flagged = [s["name"] for s in stops if s.get("boarding_hint")]
            if flagged:
                needs.append({"type": "boarding_restriction", "block": bi,
                              "stops": flagged,
                              "message": f"ブロック{bi}に乗降制約マーカー（降車専用/乗車専用）が"
                                         f"見つかりました: {' / '.join(flagged)}。対象範囲は推測して"
                                         "いません。原典で『どの停留所が乗車不可/降車不可か』を確認し、"
                                         "③で指定してください。"})
            for t in trips:
                if not t["monotonic"]:
                    needs.append({"type": "time_nonmonotonic", "block": bi, "col": t["col"],
                                  "message": f"ブロック{bi} 便(列{t['col']})で時刻が逆行しています。"
                                             "要予約への寄り道や折り返しの可能性。原典で確認してください。"})
    # 先頭のタイトル行から路線名（○○線/系統/ルート）を拾って各ブロックに付ける
    # （例:「まほろば号 湯の谷地域線」時刻表 → 湯の谷地域線）。停留所名・ファイル名で拾えない補完。
    # section_row は便ヘッダ行が無い表では None になるため、数値だけを見る
    _maxr = min([b["section_row"] for b in blocks
                 if isinstance(b.get("section_row"), int)], default=6)
    _title_texts = [ws.cell(r, c).value
                    for r in range(1, max(2, _maxr))
                    for c in range(1, ws.max_column + 1) if ws.cell(r, c).value]
    _rt = route_title_from_texts(_title_texts)
    # 運行日の表記（（平日）（土日祝）等）。タイトル→無ければシート名の順に、書かれた語だけを拾う。
    _dh = day_hint_from_texts(_title_texts) or day_hint_from_texts([ws.title])
    for b in blocks:
        b["sheet"] = ws.title                 # どのシート由来かを保持（②で見分けるため）
        if _rt:
            b["route_title"] = _rt
        if _dh:
            b["day_hint"] = _dh

    return {"blocks": blocks, "warnings": warnings_list, "needs": needs,
            "n_name_cols": len(name_cols), "n_trip_cols": len(trip_cols),
            "n_stops": len(all_served)}


def trip_signature(t):
    """便の指紋（停留所名と時刻の並び）。完全一致＝同じ便。

    冊子用シートのように同じ時刻表を別レイアウトで再掲しているブックがあり、
    そのまま取り込むと同じ便を二重に登録してしまう。停留所も時刻もすべて同じ便が
    別の便であることは実務上ありえないため、「中身が同一」という事実で重複と判定する。
    """
    return tuple((c.get("name", ""), c.get("time", "")) for c in t.get("cells", []))


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Excel時刻表から停留所・時刻を抽出（PDF座標方式と同じJSON形式で出力）")
    ap.add_argument("input", help="入力 .xlsx")
    ap.add_argument("-o", "--output", required=True, help="出力JSON")
    ap.add_argument("--sheet", default=None,
                    help="シート名（未指定なら時刻表のある全シートを読む。複数は , 区切り）")
    ap.add_argument("--name-col", default=None, help="停留所名の列（A や 3）。未指定で自動検出")
    ap.add_argument("--header-rows", type=int, default=None,
                    help="先頭の見出し行数（未指定で自動）")
    args = ap.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"Error: 入力が見つかりません: {in_path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(in_path, data_only=True)
    if args.sheet:
        want = [s.strip() for s in args.sheet.split(",") if s.strip()]
        missing = [s for s in want if s not in wb.sheetnames]
        if missing:
            print(f"Error: シートがありません: {' / '.join(missing)}", file=sys.stderr)
            return 1
        sheets = [wb[s] for s in want]
    else:
        sheets = list(wb.worksheets)        # 既定＝全シート（時刻表が無いシートは自然に除外）

    blocks, warnings_list, needs = [], [], []
    sheet_summary, skipped, seen_trip = [], [], {}
    for ws in sheets:
        res = extract_sheet(ws, args.name_col)
        warnings_list.extend(res.get("warnings", []))
        if not res["blocks"]:
            # 時刻が1つも無いシート（表紙・案内・運賃表など）は時刻表でないとみなして飛ばす
            skipped.append(ws.title)
            continue
        kept, bi_map, n_dup, n_new = [], {}, 0, 0
        for b in res["blocks"]:
            _old_bi = b.get("block_index")
            # 既に読んだ便と中身が完全に同じ便は落とす（冊子用シート等の再掲による二重登録を防ぐ）
            fresh = []
            for t in b.get("trips", []):
                sig = trip_signature(t)
                if sig in seen_trip:
                    n_dup += 1
                    continue
                seen_trip[sig] = ws.title
                fresh.append(t)
            if not fresh:
                continue
            n_new += len(fresh)
            b["trips"] = fresh
            # 残った便が通る停留所だけに絞る（落とした便にしか出てこない停留所を残さない）
            _served = {c.get("name") for t in fresh for c in t.get("cells", [])}
            b["stops"] = [s for s in b.get("stops", []) if s.get("name") in _served]
            b["block_index"] = len(blocks) + len(kept)
            bi_map[_old_bi] = b["block_index"]
            kept.append(b)
        if n_dup and not n_new:
            warnings_list.append(
                f"シート「{ws.title}」の便は、すべて他のシートと中身が完全に同じでした。"
                "二重登録を避けるため取り込んでいません（冊子用の再掲など）。")
        elif n_dup and n_new:
            # 一部だけ重複＝「再掲＋追記」か「別ダイヤ」か、機械では決められない。人に返す。
            needs.append({"type": "partial_duplicate_sheet", "sheet": ws.title,
                          "message": f"シート「{ws.title}」には他シートと同じ便が{n_dup}件、"
                                     f"同じでない便が{n_new}件ありました。重複分は取り込んでいません。"
                                     "同じ時刻表の再掲なのか、別のダイヤなのかを原典で確認してください。"})
        blocks.extend(kept)
        # ブロック番号を通し番号に振り直したので、要確認メッセージの番号も合わせる。
        # 取り込まなかった重複ブロックへの指摘はそのまま出すと番号が指す先が無いので落とす。
        for nd in res.get("needs", []):
            if "block" in nd:
                if nd["block"] not in bi_map:
                    continue
                nd = {**nd, "block": bi_map[nd["block"]]}
            nd = {**nd, "sheet": ws.title}
            needs.append(nd)
        if kept:
            sheet_summary.append({"sheet": ws.title, "blocks": len(kept),
                                  "trips": sum(len(b["trips"]) for b in kept),
                                  "stops": res.get("n_stops", 0)})
    if not blocks:
        warnings_list.append("時刻表（4:00以降の時刻セル）が見つかりませんでした。"
                             "レイアウトを確認するか --sheet でシートを指定してください。")
        needs.append({"type": "no_time_cells",
                      "message": "便の時刻セル(4:00以降のHH:MM)が無いため抽出できません。"
                                 "--sheet で正しいシートを指定するか、時刻が HH:MM 形式か確認してください。"})
    else:
        needs.append({"type": "assign_required",
                      "message": "便名・方向(direction_id)・循環の展開は表構造から確定できません。"
                                 "原典と照合して割り当ててください(Step2)。"})
    if len(sheet_summary) > 1:
        # 複数シートを1つの案件としてまとめた事実は必ず人に見せる（黙って混ぜない）。
        needs.append({"type": "multi_sheet",
                      "sheets": [s["sheet"] for s in sheet_summary],
                      "message": "複数のシートから時刻表を読み取りました（"
                                 + " / ".join(f"{s['sheet']}:{s['trips']}便" for s in sheet_summary)
                                 + "）。②でシートごとに路線名と運行する曜日を割り当ててください。"})

    result = {"source": str(in_path),
              "sheet": sheet_summary[0]["sheet"] if len(sheet_summary) == 1 else "",
              "sheets": [s["sheet"] for s in sheet_summary],
              "sheet_summary": sheet_summary, "skipped_sheets": skipped,
              "blocks": blocks, "warnings": warnings_list,
              "needs_confirmation": needs}
    Path(args.output).write_text(json.dumps(result, ensure_ascii=False, indent=2),
                                 encoding="utf-8")

    # サマリ（cp932安全・絵文字なし）
    total_trips = sum(len(b["trips"]) for b in blocks)
    for s in sheet_summary:
        print(f"[INFO] シート: {s['sheet']}  表={s['blocks']}  便={s['trips']}", file=sys.stderr)
    if skipped:
        print(f"[INFO] 時刻表なしで除外: {' / '.join(skipped)}", file=sys.stderr)
    print(f"[INFO] 合計 表(領域×セクション)={len(blocks)} / 便={total_trips}", file=sys.stderr)
    if needs:
        print(f"[INFO] 要確認: {len(needs)}件", file=sys.stderr)
    print(f"[OK] 出力: {args.output}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
