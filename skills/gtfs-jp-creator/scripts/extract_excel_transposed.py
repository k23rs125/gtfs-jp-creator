# -*- coding: utf-8 -*-
"""転置Excel時刻表（停留所が列・便が行）→ 抽出JSON（座標方式/Excel直接と同形式）。

extract_timetable_excel.py は「停留所が行・便が列」専用。本スクリプトはその逆向き
（停留所が列・便が行）に対応する。各シートを1ブロック（=1方向）として扱う。
JR到着/発車時刻・市営バス接続・所要時間など、停留所でない列は見出しのパターンで除外する。

判断（路線・方向・循環）は行わない（Step2のLLM/利用者に委ねる）。

Usage:
  python extract_excel_transposed.py <xlsx> -o out.json [--header-row 2]
      [--exclude "時刻|着時|発時|到着|発車|所要"] [--all-sheets]
"""
import argparse, json, re, unicodedata
import openpyxl
from datetime import time as dtime

def norm(s):
    if s is None: return ""
    s = unicodedata.normalize("NFKC", str(s)).replace("　", " ")
    return "".join(s.split())   # 内部空白を全除去（「遠賀川駅 前」→「遠賀川駅前」）

def to_hhmmss(v):
    if isinstance(v, dtime):
        return f"{v.hour:02d}:{v.minute:02d}:{v.second:02d}"
    if isinstance(v, str):
        m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$", v.strip())
        if m:
            return f"{int(m.group(1)):02d}:{m.group(2)}:{m.group(3) or '00'}"
    return None

def extract_sheet(ws, header_row, exclude_re):
    # 停留所列＝見出し行に名前があり、除外パターンに当たらない列
    stop_cols = []
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, c).value
        nm = norm(raw)
        if nm and not exclude_re.search(nm):
            stop_cols.append((c, nm))
    # 便番号列の推定（見出し行が空で、下の行に小整数が並ぶ列）
    # 便（行）＝停留所列のどれかに時刻がある行
    trips = []
    for r in range(header_row + 1, ws.max_row + 1):
        cells = []
        seq = 0
        for c, nm in stop_cols:
            t = to_hhmmss(ws.cell(r, c).value)
            if t:
                seq += 1
                cells.append({"seq": seq, "num": None, "name": nm, "time": t, "reserve": False})
        if len(cells) >= 2:
            trips.append({"row": r, "trip_number": len(trips) + 1, "cells": cells})
    return stop_cols, trips

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input")
    ap.add_argument("-o", "--output", required=True)
    ap.add_argument("--header-row", type=int, default=2)
    ap.add_argument("--exclude", default="時刻|着時|発時|到着|発車|所要|運賃|備考")
    ap.add_argument("--all-sheets", action="store_true", default=True)
    ap.add_argument("--sheet", default=None)
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.input, data_only=True)
    exclude_re = re.compile(a.exclude)
    sheets = [wb[a.sheet]] if a.sheet else wb.worksheets

    blocks = []
    for bi, ws in enumerate(sheets):
        stop_cols, trips = extract_sheet(ws, a.header_row, exclude_re)
        blocks.append({
            "block_index": bi,
            "sheet": ws.title.strip(),
            "stops": [{"name": nm} for _, nm in stop_cols],
            "trips": trips,
        })
        print(f"[block {bi}] {ws.title.strip()}: 停留所{len(stop_cols)} 便{len(trips)}")

    out = {
        "source": a.input,
        "layout": "transposed (stops=columns, trips=rows)",
        "blocks": blocks,
        "warnings": [],
        "needs_confirmation": [
            {"type": "assign_required",
             "message": "便名(A1便等)・方向(direction_id)・循環の展開は座標から確定できません。"
                        "原典と照合して割り当ててください。各シートが1方向に対応する想定。"}
        ],
    }
    json.dump(out, open(a.output, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"[OK] 出力: {a.output}")

if __name__ == "__main__":
    main()
